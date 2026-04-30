# routes/advanced.py — Blueprint: advanced

from flask import Blueprint, request, render_template, redirect, url_for, send_file
from werkzeug.utils import secure_filename
from utils import parsear_paginas, hex_a_rgb
from config import UPLOAD_FOLDER, OUTPUT_FOLDER
from auth import login_required
import fitz
import os
import io
import zipfile
import pdfplumber
from openpyxl import Workbook
import pytesseract
from PIL import Image
import difflib


advanced_bp = Blueprint('advanced', __name__)


@advanced_bp.route('/sign_ui')
@login_required
def sign_ui():
    """Renderiza la interfaz interactiva para firmar PDFs."""
    return render_template('sign.html')


@advanced_bp.route('/form_filler', methods=['POST'])
@login_required
def form_filler():
    """
    Paso 1 del rellenador de formularios: detecta los campos del PDF.

    Lee todos los widgets (campos interactivos) del PDF y los pasa
    al template form_filler.html donde el usuario los puede llenar.
    El nombre del archivo se pasa como campo oculto para el paso 2.

    Parámetros del formulario:
        pdf_file (file): PDF con formulario interactivo.

    Returns:
        Response: Template form_filler.html con los campos detectados.

    Tipos de campo soportados:
        PDF_WIDGET_TYPE_TEXT: campos de texto
        PDF_WIDGET_TYPE_CHECKBOX: casillas de verificación
        PDF_WIDGET_TYPE_LISTBOX: listas desplegables
    
    Cada campo incluye:
        - nombre: nombre del campo
        - tipo: tipo de widget
        - valor_actual: valor actual del campo
        - pagina: número de página (base 0)
        - rect: diccionario con coordenadas del campo (x0, y0, x1, y1, page_width, page_height)
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    try:
        doc = fitz.open(pdf_path)
        campos = []

        for page_num, page in enumerate(doc):
            for field in page.widgets():
                if field.field_type in [
                    fitz.PDF_WIDGET_TYPE_TEXT,
                    fitz.PDF_WIDGET_TYPE_CHECKBOX,
                    fitz.PDF_WIDGET_TYPE_LISTBOX
                ]:
                    campos.append({
                        'nombre': field.field_name,
                        'tipo': field.field_type,
                        'valor_actual': field.field_value or '',
                        'pagina': page_num,
                        'rect': {
                            'x0': field.rect.x0,
                            'y0': field.rect.y0,
                            'x1': field.rect.x1,
                            'y1': field.rect.y1,
                            'page_width': page.rect.width,
                            'page_height': page.rect.height
                        }
                    })

        doc.close()

        if not campos:
            return 'Este PDF no contiene campos de formulario.', 400

    except Exception as e:
        return f'Error al leer el formulario: {str(e)}', 500

    return render_template('form_filler.html', campos=campos, pdf_nombre=file.filename)


@advanced_bp.route('/form_filler/guardar', methods=['POST'])
@login_required
def form_filler_guardar():
    """
    Paso 2 del rellenador de formularios: escribe los valores en el PDF.

    Recibe el nombre del archivo original (pasado como campo oculto desde
    form_filler.html) y los valores de cada campo del formulario.
    Escribe cada valor en su campo correspondiente y guarda el resultado.

    Parámetros del formulario:
        pdf_nombre (str): Nombre del PDF original en /uploads.
        [campo_name] (str/on): Valor de cada campo del formulario.

    Returns:
        Response: Template con enlace al PDF rellenado.

    Validaciones:
        - pdf_nombre vacío -> 400
        - Archivo original no encontrado en /uploads -> 400
    """
    pdf_nombre = request.form.get('pdf_nombre', '')

    if not pdf_nombre:
        return 'Nombre de archivo no encontrado.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, pdf_nombre)

    if not os.path.exists(pdf_path):
        return 'El archivo original ya no está disponible. Sube el PDF nuevamente.', 400

    output_filename = pdf_nombre.rsplit('.', 1)[0] + '_rellenado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        for page in doc:
            for field in page.widgets():
                nombre = field.field_name
                valor = request.form.get(nombre, '')

                # Los checkboxes envían 'on' cuando están marcados, nada cuando no
                if field.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                    field.field_value = valor == 'on'
                else:
                    field.field_value = valor

                field.update()  # aplicar el cambio al PDF en memoria

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al guardar el formulario: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@advanced_bp.route('/sign', methods=['POST'])
@login_required
def sign_pdf():
    """
    Inserta una imagen de firma en una página específica del PDF.

    La firma no es criptográfica — es una imagen visible insertada en
    el área que el usuario defina. La posición y tamaño se expresan
    como porcentaje de las dimensiones de la página.

    Parámetros del formulario:
        pdf_file (file): PDF a firmar.
        firma_file (file): Imagen de la firma (JPG o PNG).
        pagina (int): Número de página (base 1). Por defecto: 1.
        pos_x (int): Posición horizontal en % (0-90). Por defecto: 60.
        pos_y (int): Posición vertical en % (0-90). Por defecto: 80.
        firma_ancho (int): Ancho de la firma en % (5-50). Por defecto: 30.
        firma_alto (int): Alto de la firma en % (5-30). Por defecto: 10.

    Returns:
        Response: Template con enlace al PDF firmado.
    """
    if 'pdf_file' not in request.files or 'firma_file' not in request.files:
        return 'Por favor, sube el PDF y la imagen de tu firma.', 400

    file = request.files['pdf_file']
    firma = request.files['firma_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF válido.', 400

    extensiones_firma = {'.jpg', '.jpeg', '.png'}
    ext_firma = os.path.splitext(firma.filename)[1].lower()
    if ext_firma not in extensiones_firma:
        return 'La firma debe ser una imagen JPG o PNG.', 400

    try:
        pagina_num = max(1, int(request.form.get('pagina', 1)))
        pos_x = max(0, min(90, int(request.form.get('pos_x', 60))))
        pos_y = max(0, min(90, int(request.form.get('pos_y', 80))))
        firma_ancho = max(5, min(50, int(request.form.get('firma_ancho', 30))))
        firma_alto = max(5, min(30, int(request.form.get('firma_alto', 10))))
    except ValueError:
        return 'Los valores de posición deben ser números enteros.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_firmado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        # La validación de página ocurre DESPUÉS de abrir el PDF
        # porque necesitamos page_count para saber si la página existe
        if pagina_num > doc.page_count:
            doc.close()
            return f'El PDF solo tiene {doc.page_count} páginas.', 400

        page = doc[pagina_num - 1]  # convertir de base 1 a base 0
        ancho = page.rect.width
        alto = page.rect.height

        # Calcular el rectángulo donde se insertará la firma
        x0 = ancho * (pos_x / 100)
        y0 = alto * (pos_y / 100)
        x1 = x0 + ancho * (firma_ancho / 100)
        y1 = y0 + alto * (firma_alto / 100)

        firma_bytes = firma.read()
        page.insert_image(fitz.Rect(x0, y0, x1, y1), stream=firma_bytes)

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al firmar el documento: {str(e)}', 500

    return send_file(
        output_path,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=output_filename
    )


@advanced_bp.route('/pdf_to_excel', methods=['POST'])
@login_required
def pdf_to_excel():
    """
    Extrae todas las tablas de un PDF y las exporta a un archivo Excel.

    Cada tabla encontrada se escribe en una hoja separada. El nombre de
    cada hoja indica la página y número de tabla de origen (ej: Pag1_Tabla2).
    Las celdas vacías (None) se convierten a string vacío para evitar
    corrupción del archivo Excel.

    Parámetros del formulario:
        pdf_file (file): PDF con tablas a extraer.

    Returns:
        Response: Template con enlace al archivo Excel generado.

    Validaciones:
        - PDF sin tablas detectables -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '.xlsx'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Workbook() crea una hoja vacía por defecto — la eliminamos
        tablas_encontradas = 0

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tablas = page.extract_tables()

                for tabla_num, tabla in enumerate(tablas):
                    if not tabla:
                        continue

                    tablas_encontradas += 1
                    nombre_hoja = f"Pag{page_num + 1}_Tabla{tabla_num + 1}"
                    ws = wb.create_sheet(title=nombre_hoja)

                    for fila in tabla:
                        # pdfplumber retorna None en celdas vacías — openpyxl necesita strings
                        fila_limpia = [celda if celda is not None else '' for celda in fila]
                        ws.append(fila_limpia)

        if tablas_encontradas == 0:
            return 'No se encontraron tablas en el PDF.', 400

        wb.save(output_path)

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@advanced_bp.route('/edit', methods=['POST'])
@login_required
def edit_pdf():
    """
    Agrega texto en una posición específica de una página del PDF.

    Soporta dos modos:
    1. Modo legacy (backward compatible): un solo texto con campos individuales
    2. Modo nuevo: JSON array de anotaciones via 'anotaciones' field

    Parámetros del formulario (legacy):
        pdf_file (file): PDF a editar.
        texto (str): Texto a insertar.
        pagina (int): Número de página (base 1). Por defecto: 1.
        pos_x (int): Posición horizontal en % (0-95). Por defecto: 10.
        pos_y (int): Posición vertical en % (0-95). Por defecto: 50.
        fontsize (int): Tamaño de fuente en puntos (6-72). Por defecto: 12.
        color (str): Color del texto en hex. Por defecto: '#000000'.

    Parámetros del formulario (nuevo):
        pdf_file (file): PDF a editar.
        anotaciones (str): JSON array de anotaciones:
            [{"pagina": 1, "texto": "...", "pos_x": 10, "pos_y": 50, 
              "fontsize": 12, "color": "#ff0000"}, ...]

    Returns:
        Response: Template con enlace al PDF editado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_editado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        anotaciones_json = request.form.get('anotaciones')
        
        if anotaciones_json:
            import json
            try:
                anotaciones = json.loads(anotaciones_json)
                for a in anotaciones:
                    pagina_num = max(1, int(a.get('pagina', 1)))
                    texto = a.get('texto', '').strip()
                    pos_x = max(0, min(95, int(a.get('pos_x', 10))))
                    pos_y = max(0, min(95, int(a.get('pos_y', 50))))
                    fontsize = max(6, min(72, int(a.get('fontsize', 12))))
                    color_hex = a.get('color', '#000000')

                    if pagina_num > doc.page_count or not texto:
                        continue

                    page = doc[pagina_num - 1]
                    ancho = page.rect.width
                    alto = page.rect.height

                    x = ancho * (pos_x / 100)
                    y = alto * (pos_y / 100)

                    color = hex_a_rgb(color_hex)
                    page.insert_text(fitz.Point(x, y), texto, fontsize=fontsize, color=color)
            except (json.JSONDecodeError, ValueError) as e:
                doc.close()
                return f'Error al procesar anotaciones: {str(e)}', 400
        else:
            texto = request.form.get('texto', '').strip()
            color_hex = request.form.get('color', '#000000')

            if not texto:
                doc.close()
                return 'Por favor, escribe el texto a insertar.', 400

            try:
                pagina_num = max(1, int(request.form.get('pagina', 1)))
                pos_x = max(0, min(95, int(request.form.get('pos_x', 10))))
                pos_y = max(0, min(95, int(request.form.get('pos_y', 50))))
                fontsize = max(6, min(72, int(request.form.get('fontsize', 12))))
            except ValueError:
                doc.close()
                return 'Los valores de posición deben ser números enteros.', 400

            if pagina_num > doc.page_count:
                doc.close()
                return f'El PDF solo tiene {doc.page_count} páginas.', 400

            page = doc[pagina_num - 1]
            ancho = page.rect.width
            alto = page.rect.height

            x = ancho * (pos_x / 100)
            y = alto * (pos_y / 100)

            color = hex_a_rgb(color_hex)
            page.insert_text(fitz.Point(x, y), texto, fontsize=fontsize, color=color)

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al editar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')