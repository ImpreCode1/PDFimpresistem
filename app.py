"""
PDFimpresistem - Aplicación Flask para manipulación de archivos PDF
====================================================================
Empresa: Impresistem S.A.S - Cota, Cundinamarca
Rama: developer

Descripción:
    Aplicación web que centraliza las operaciones más comunes sobre
    documentos PDF. El usuario sube un archivo desde el navegador,
    el servidor lo procesa y retorna el resultado para descarga.

Patrón base de cada ruta:
    1. Validar que el archivo fue enviado (request.files)
    2. Validar extensión .pdf
    3. Guardar en /uploads con secure_filename
    4. Procesar con la librería correspondiente (try/except)
    5. Guardar resultado en /outputs
    6. Retornar render_template con output_file o send_file

Librerías principales:
    - PyMuPDF (fitz): operaciones sobre páginas PDF
    - pikepdf: conversión a PDF/A
    - pdfplumber: extracción de tablas
    - pytesseract: reconocimiento óptico de caracteres (OCR)
    - APScheduler: limpieza automática de archivos a las 7 PM Colombia
"""

# ─── Imports ──────────────────────────────────────────────────────────────────

from flask import Flask, request, render_template, send_from_directory, send_file
from pdf2docx import Converter
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from PIL import Image
import pytz
import shutil
import atexit
import fitz          # PyMuPDF
import io
import os
import zipfile
import pikepdf
import pdfplumber
import pytesseract
import difflib

# ─── Configuración de la aplicación ───────────────────────────────────────────

app = Flask(__name__)

# Rutas de carpetas — se pueden sobreescribir con variables de entorno
# en el servidor de producción sin modificar el código
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'uploads'))
OUTPUT_FOLDER = os.getenv('OUTPUT_FOLDER', os.path.join(BASE_DIR, 'outputs'))

# Crear carpetas si no existen (útil en primer despliegue)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ─── Limpieza automática de archivos ──────────────────────────────────────────

def limpiar_carpeta(carpeta):
    """
    Elimina todos los archivos y subcarpetas dentro de una carpeta
    sin borrar la carpeta misma.

    Si un archivo está en uso y no puede eliminarse, registra el error
    en consola y continúa con el siguiente archivo sin interrumpir el proceso.

    Args:
        carpeta (str): Ruta absoluta de la carpeta a limpiar.
    """
    for nombre in os.listdir(carpeta):
        ruta = os.path.join(carpeta, nombre)
        try:
            if os.path.isfile(ruta):
                os.remove(ruta)
            elif os.path.isdir(ruta):
                shutil.rmtree(ruta)
        except Exception as e:
            print(f'[Limpieza] Error eliminando {ruta}: {e}')


def limpiar_archivos_programada():
    """
    Tarea programada que limpia /uploads y /outputs a las 7 PM hora Colombia.
    Es ejecutada automáticamente por APScheduler según el CronTrigger configurado.
    """
    print('[Limpieza] Ejecutando limpieza programada...')
    limpiar_carpeta(UPLOAD_FOLDER)
    limpiar_carpeta(OUTPUT_FOLDER)
    print('[Limpieza] Limpieza completada.')


# Configurar el scheduler con zona horaria de Colombia (UTC-5)
zona_colombia = pytz.timezone('America/Bogota')

scheduler = BackgroundScheduler(timezone=zona_colombia)
scheduler.add_job(
    limpiar_archivos_programada,
    CronTrigger(hour=19, minute=0, timezone=zona_colombia)  # 7:00 PM Colombia
)
scheduler.start()

# Garantizar cierre limpio del scheduler al detener el servidor con Ctrl+C
# Sin esto, el hilo del scheduler quedaría corriendo en segundo plano
atexit.register(lambda: scheduler.shutdown(wait=False))

# ─── Funciones auxiliares ─────────────────────────────────────────────────────

def parsear_paginas(texto, total_paginas):
    """
    Convierte un string de páginas a una lista de índices base 0.

    Acepta formatos como:
        - Páginas individuales: "1,3,5"
        - Rangos: "2-6"
        - Combinaciones: "1,3-5,8"

    Usa set() internamente para eliminar duplicados automáticamente.
    Filtra silenciosamente las páginas fuera del rango del documento.

    Args:
        texto (str): String con páginas escritas por el usuario.
        total_paginas (int): Total de páginas del documento.

    Returns:
        list[int]: Lista ordenada de índices base 0.

    Raises:
        ValueError: Si el formato de un rango es inválido (ej: "5-3").
    """
    paginas = set()
    partes = texto.split(',')
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        if '-' in parte:
            extremos = parte.split('-', 1)
            if len(extremos) != 2 or not extremos[0].strip() or not extremos[1].strip():
                raise ValueError('Formato de rango inválido.')
            inicio = int(extremos[0])
            fin = int(extremos[1])
            if inicio > fin:
                raise ValueError('El rango de páginas debe ir de menor a mayor.')
            for n in range(inicio, fin + 1):
                paginas.add(n - 1)
        else:
            paginas.add(int(parte) - 1)
    return sorted([p for p in paginas if 0 <= p < total_paginas])


def hex_a_rgb(hex_color):
    """
    Convierte un color hexadecimal HTML a una tupla RGB normalizada (0.0 - 1.0).

    PyMuPDF requiere colores en formato RGB con valores entre 0 y 1,
    mientras que los inputs de tipo 'color' en HTML retornan hex como '#ff0000'.

    Ejemplo:
        hex_a_rgb('#ff0000') -> (1.0, 0.0, 0.0)  # rojo
        hex_a_rgb('#000000') -> (0.0, 0.0, 0.0)  # negro

    Args:
        hex_color (str): Color en formato hexadecimal, ej: '#ff0000' o 'ff0000'.

    Returns:
        tuple[float, float, float]: Tupla (r, g, b) con valores entre 0.0 y 1.0.
    """
    hex_color = hex_color.lstrip('#')
    r = int(hex_color[0:2], 16) / 255  # int('ff', 16) = 255 / 255 = 1.0
    g = int(hex_color[2:4], 16) / 255
    b = int(hex_color[4:6], 16) / 255
    return (r, g, b)

# ─── Rutas principales ────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Renderiza la página principal con todas las tarjetas de funciones."""
    return render_template('index.html', output_file=None)


@app.route('/cerrar_sesion', methods=['POST'])
def cerrar_sesion():
    """
    Limpieza manual de archivos activada por el usuario.
    El botón 'Limpiar archivos' del header apunta a esta ruta.
    Elimina todos los archivos en /uploads y /outputs y redirige al inicio.
    """
    limpiar_carpeta(UPLOAD_FOLDER)
    limpiar_carpeta(OUTPUT_FOLDER)
    return render_template('index.html', output_file=None)


@app.route('/download/<filename>')
def download_file(filename):
    """
    Sirve un archivo desde la carpeta /outputs para descarga.

    Args:
        filename (str): Nombre del archivo a descargar.

    Returns:
        Response: Archivo como adjunto descargable.
    """
    return send_from_directory(OUTPUT_FOLDER, filename)

# ─── Nivel Básico ─────────────────────────────────────────────────────────────

@app.route('/rotate', methods=['POST'])
def rotate_pdf():
    """
    Rota todas las páginas de un PDF en el ángulo indicado.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a rotar.
        angulo (str): Ángulo de rotación. Debe ser '90', '180' o '270'.

    Returns:
        Response: Template con enlace al PDF rotado.

    Validaciones:
        - Archivo no seleccionado -> 400
        - Extensión diferente a .pdf -> 400
        - Ángulo no es 90/180/270 -> 400
        - Ángulo no numérico -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    angulo = request.form.get('angulo', '90')

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    try:
        angulo = int(angulo)
        if angulo not in [90, 180, 270]:
            return 'Ángulo inválido. Use 90, 180 o 270.', 400
    except ValueError:
        return 'El ángulo debe ser un número.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_rotado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    # set_rotation() aplica la rotación a cada página individualmente
    doc = fitz.open(pdf_path)
    for page in doc:
        page.set_rotation(angulo)
    doc.save(output_path)
    doc.close()

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/extract', methods=['POST'])
def extract_pages():
    """
    Extrae páginas específicas de un PDF y las guarda en un nuevo documento.

    El usuario puede especificar páginas individuales (1,3,5), rangos (2-6)
    o combinaciones (1,3-5,8). Usa parsear_paginas() para procesar el input.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF fuente.
        paginas (str): Páginas a extraer. Ej: '1,3-5,8'.

    Returns:
        Response: Template con enlace al PDF extraído.

    Validaciones:
        - Archivo no seleccionado -> 400
        - Campo de páginas vacío -> 400
        - Páginas fuera de rango -> se filtran silenciosamente
        - Lista vacía tras filtro -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    paginas_texto = request.form.get('paginas', '').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not paginas_texto:
        return 'Por favor, indica las páginas a extraer.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_extraido.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc_original = fitz.open(pdf_path)
        total = doc_original.page_count
        indices = parsear_paginas(paginas_texto, total)

        if not indices:
            return 'No se encontraron páginas válidas en el rango indicado.', 400

        # Crear documento nuevo e insertar solo las páginas seleccionadas
        doc_nuevo = fitz.open()
        for i in indices:
            doc_nuevo.insert_pdf(doc_original, from_page=i, to_page=i)

        doc_nuevo.save(output_path)
        doc_nuevo.close()
        doc_original.close()

    except ValueError as e:
        return f'Formato de páginas inválido: {str(e)}', 400
    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/watermark', methods=['POST'])
def watermark_pdf():
    """
    Inserta texto semitransparente en diagonal sobre todas las páginas del PDF.

    El texto se centra dinámicamente en cada página según sus dimensiones,
    con rotación de 45° y opacidad del 30% para no ocultar el contenido.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a marcar.
        texto (str): Texto de la marca de agua. Por defecto: 'CONFIDENCIAL'.

    Returns:
        Response: Template con enlace al PDF con marca de agua.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    texto = request.form.get('texto', 'CONFIDENCIAL').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not texto:
        return 'Por favor, escribe el texto de la marca de agua.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_marcado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        for page in doc:
            ancho = page.rect.width
            alto = page.rect.height

            # El centro se calcula por página porque pueden tener distintos tamaños
            centro_x = ancho / 2
            centro_y = alto / 2

            page.insert_text(
                fitz.Point(centro_x - 150, centro_y),
                texto,
                fontsize=60,
                color=(0.6, 0.6, 0.6),   # gris en formato RGB normalizado
                rotate=45,                 # diagonal
                fill_opacity=0.3           # 30% de opacidad — semitransparente
            )

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/protect', methods=['POST'])
def protect_pdf():
    """
    Encripta un PDF con contraseña usando el algoritmo AES-256.

    Genera automáticamente un owner_pwd diferente al user_pwd para
    evitar que quien tenga la contraseña de apertura pueda editar
    los permisos del documento.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a proteger.
        password (str): Contraseña para abrir el documento.
        confirm_password (str): Confirmación de la contraseña.

    Returns:
        Response: Template con enlace al PDF protegido.

    Validaciones:
        - Contraseña vacía -> 400
        - Contraseñas no coinciden -> 400
        - Contraseña menor a 4 caracteres -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    password = request.form.get('password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not password:
        return 'Por favor, ingresa una contraseña.', 400

    if password != confirm_password:
        return 'Las contraseñas no coinciden.', 400

    if len(password) < 4:
        return 'La contraseña debe tener al menos 4 caracteres.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_protegido.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)
        doc.save(
            output_path,
            user_pwd=password,                    # contraseña para abrir
            owner_pwd=password + "_owner",         # contraseña de permisos (oculta al usuario)
            encryption=fitz.PDF_ENCRYPT_AES_256   # estándar de encriptación actual
        )
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/unlock', methods=['POST'])
def unlock_pdf():
    """
    Elimina la protección por contraseña de un PDF encriptado.

    Verifica primero que el PDF esté encriptado antes de intentar
    autenticar. Guarda sin parámetros de contraseña para quitar la protección.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF protegido.
        password (str): Contraseña del documento.

    Returns:
        Response: Template con enlace al PDF desbloqueado.

    Validaciones:
        - PDF no está encriptado -> 400
        - Contraseña incorrecta -> 400 (authenticate() retorna 0)
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    password = request.form.get('password', '').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not password:
        return 'Por favor, ingresa la contraseña del documento.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_desbloqueado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        # Verificar antes de intentar autenticar — evita comportamiento inesperado
        if not doc.is_encrypted:
            doc.close()
            return 'Este PDF no está protegido con contraseña.', 400

        # authenticate() retorna 0 si la contraseña es incorrecta
        resultado = doc.authenticate(password)
        if resultado == 0:
            doc.close()
            return 'Contraseña incorrecta.', 400

        # Guardar sin user_pwd ni owner_pwd elimina la protección
        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/flatten', methods=['POST'])
def flatten_pdf():
    """
    Aplana un PDF convirtiendo anotaciones y campos de formulario en contenido estático.

    El resultado no puede ser editado — los campos interactivos quedan
    como texto plano incrustado en la página. Útil para enviar formularios
    diligenciados sin posibilidad de modificación posterior.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF con formularios o anotaciones.

    Returns:
        Response: Template con enlace al PDF aplanado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_aplanado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        for page in doc:
            # Paso 1: aplicar visualmente las anotaciones al contenido base
            annots = page.annots()
            if annots:
                for annot in annots:
                    annot.update()

            # Paso 2: eliminar los campos de formulario interactivos
            widgets = page.widgets()
            if widgets:
                for field in widgets:
                    page.delete_widget(field)

            # Paso 3: fusionar y limpiar las capas de la página
            page.clean_contents()

        # deflate=True compensa el posible aumento de tamaño al fusionar capas
        doc.save(output_path, deflate=True)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/reorder', methods=['POST'])
def reorder_pdf():
    """
    Reorganiza las páginas de un PDF en el orden indicado por el usuario.

    A diferencia de extract_pages(), requiere que estén TODAS las páginas
    del documento exactamente una vez. Usa doc.select() que es más eficiente
    que crear un documento nuevo.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a reordenar.
        paginas (str): Nuevo orden de páginas. Ej: '3,1,2' para un PDF de 3 páginas.

    Returns:
        Response: Template con enlace al PDF reordenado.

    Validaciones:
        - Páginas repetidas -> 400 (detectado por len != total)
        - Páginas faltantes -> 400 (detectado por sorted != range)
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    paginas_texto = request.form.get('paginas', '').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not paginas_texto:
        return 'Por favor, indica el nuevo orden de las páginas.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_ordenado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)
        total = doc.page_count
        indices = parsear_paginas(paginas_texto, total)

        # Validación doble: detecta páginas faltantes Y páginas repetidas en un paso
        # - len != total: faltan páginas (set() eliminó duplicados)
        # - sorted != range: hay páginas fuera del rango o el conteo no cierra
        if len(indices) != total or sorted(indices) != list(range(total)):
            doc.close()
            return f'Debes incluir todas las {total} páginas exactamente una vez.', 400

        doc.select(indices)  # reordena en memoria eficientemente
        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/organize', methods=['POST'])
def organize_pdf():
    """
    Elimina páginas específicas de un PDF.

    Lógica inversa a extract_pages(): el usuario indica qué páginas ELIMINAR
    en vez de cuáles conservar. Se construye la lista de páginas a conservar
    como el complemento de las páginas a eliminar.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a organizar.
        paginas (str): Páginas a eliminar. Ej: '2,5,8' o '3-6'.

    Returns:
        Response: Template con enlace al PDF organizado.

    Validaciones:
        - Intento de eliminar todas las páginas -> 400
        - Páginas fuera de rango -> se ignoran silenciosamente
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    paginas_texto = request.form.get('paginas', '').strip()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not paginas_texto:
        return 'Por favor, indica las páginas a eliminar.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_organizado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)
        total = doc.page_count

        indices_eliminar = set(parsear_paginas(paginas_texto, total))

        # Un PDF con 0 páginas es un archivo corrupto
        if len(indices_eliminar) >= total:
            doc.close()
            return 'No puedes eliminar todas las páginas del documento.', 400

        # Lógica inversa: conservar todo lo que NO está en la lista a eliminar
        indices_conservar = [i for i in range(total) if i not in indices_eliminar]

        doc.select(indices_conservar)
        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/page_numbers', methods=['POST'])
def page_numbers_pdf():
    """
    Inserta numeración automática en el pie de cada página del PDF.

    El texto tiene el formato 'Página X de Y'. La posición horizontal
    se compensa con un offset porque insert_text() dibuja desde el punto
    hacia la derecha, no centrado en él.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a numerar.
        posicion (str): 'centro', 'derecha' o 'izquierda'. Por defecto: 'centro'.

    Returns:
        Response: Template con enlace al PDF numerado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    posicion = request.form.get('posicion', 'centro')

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_numerado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)
        total = doc.page_count

        for i, page in enumerate(doc):
            numero = i + 1  # el usuario ve páginas desde 1, no desde 0
            ancho = page.rect.width
            alto = page.rect.height
            texto = f"Página {numero} de {total}"

            # Coordenada Y siempre al pie — 20 puntos desde el borde inferior
            # Coordenada X varía según posición elegida
            if posicion == 'centro':
                x = ancho / 2 - 30   # -30 compensa el ancho del texto
                y = alto - 20
            elif posicion == 'derecha':
                x = ancho - 80        # -80 para que el texto no se corte
                y = alto - 20
            elif posicion == 'izquierda':
                x = 20
                y = alto - 20
            else:
                x = ancho / 2 - 30   # valor por defecto: centro
                y = alto - 20

            page.insert_text(fitz.Point(x, y), texto, fontsize=10, color=(0, 0, 0))

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')

# ─── Nivel Intermedio ─────────────────────────────────────────────────────────

@app.route('/compress', methods=['POST'])
def compress_pdf():
    """
    Reduce el tamaño en disco de un PDF optimizando su estructura interna.

    Calcula el ahorro en porcentaje comparando tamaños antes y después.
    Si el archivo ya estaba optimizado, el ahorro puede ser 0% o negativo
    (en cuyo caso se muestra 0% para no confundir al usuario).

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a comprimir.

    Returns:
        Response: Template con enlace al PDF comprimido y porcentaje de ahorro.

    Parámetros de save():
        garbage=4: limpieza máxima de recursos duplicados y no usados
        deflate=True: compresión de streams de contenido
        clean=True: optimización de la estructura interna
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_comprimido.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)
        doc.save(output_path, garbage=4, deflate=True, clean=True)
        doc.close()  # cerrar ANTES de medir — garantiza escritura completa en disco

        tamaño_original = os.path.getsize(pdf_path)
        tamaño_comprimido = os.path.getsize(output_path)
        ahorro = round((1 - tamaño_comprimido / tamaño_original) * 100, 1)

        if ahorro < 0:
            ahorro = 0  # archivo ya estaba optimizado

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template(
        'index.html',
        output_file=f'/download/{output_filename}',
        ahorro=ahorro   # variable extra pasada al template para mostrar el porcentaje
    )


@app.route('/pdf_to_jpg', methods=['POST'])
def pdf_to_jpg():
    """
    Convierte cada página del PDF en una imagen y las empaqueta en un ZIP.

    Usa BytesIO para crear el ZIP en memoria sin escribir archivos temporales
    en disco. El resultado se envía directamente con send_file() sin recargar
    la página (no usa render_template).

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a convertir.
        dpi (str): Resolución de las imágenes. Opciones: 72, 150, 300. Por defecto: 150.
        image_format (str): Formato de salida. 'jpg' o 'png'. Por defecto: 'png'.

    Returns:
        Response: Archivo ZIP descargable con las imágenes.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    dpi = request.form.get('dpi', '150')
    image_format = request.form.get('image_format', 'png').lower()

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    try:
        dpi = int(dpi)
        if dpi not in [72, 150, 300]:
            dpi = 150
    except ValueError:
        dpi = 150

    if image_format not in {'jpg', 'png'}:
        image_format = 'png'

    pdf_path = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))
    file.save(pdf_path)

    nombre_base = file.filename.rsplit('.', 1)[0]

    try:
        doc = fitz.open(pdf_path)
        buffer = io.BytesIO()  # ZIP en memoria — no toca el disco

        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for i, page in enumerate(doc):
                pixmap = page.get_pixmap(dpi=dpi)
                formato_pymupdf = 'jpg' if image_format == 'jpg' else 'png'
                extension = 'jpg' if image_format == 'jpg' else 'png'
                img_bytes = pixmap.tobytes(formato_pymupdf)
                nombre_imagen = f"{nombre_base}_pagina_{i + 1}.{extension}"
                zipf.writestr(nombre_imagen, img_bytes)  # escribe bytes directo al ZIP

        doc.close()
        buffer.seek(0)  # volver al inicio del buffer para que send_file lo lea

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return send_file(
        buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{nombre_base}_imagenes_{image_format}.zip'
    )


@app.route('/jpg_to_pdf', methods=['POST'])
def jpg_to_pdf():
    """
    Convierte una o varias imágenes JPG/PNG en un único PDF.

    Usa request.files.getlist() para recibir múltiples archivos. Crea cada
    página del tamaño exacto de la imagen leyendo sus dimensiones con fitz
    antes de insertarla, evitando distorsión o bordes blancos.

    Parámetros del formulario:
        imagenes (file[]): Una o varias imágenes JPG o PNG.

    Returns:
        Response: Template con enlace al PDF generado.

    Validaciones:
        - Sin archivos -> 400
        - Formato no soportado -> 400 indicando el archivo específico
    """
    if 'imagenes' not in request.files:
        return 'No se han seleccionado imágenes.', 400

    imagenes = request.files.getlist('imagenes')
    extensiones_validas = {'.jpg', '.jpeg', '.png'}

    if not imagenes or imagenes[0].filename == '':
        return 'Por favor, selecciona al menos una imagen.', 400

    for img in imagenes:
        ext = os.path.splitext(img.filename)[1].lower()
        if ext not in extensiones_validas:
            return f'Formato no soportado: {img.filename}. Use JPG o PNG.', 400

    output_filename = 'imagenes_convertidas.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open()

        for imagen in imagenes:
            imagen_bytes = imagen.read()
            ext = os.path.splitext(imagen.filename)[1].lower()
            filetype = 'png' if ext == '.png' else 'jpeg'

            # Abrir temporalmente para leer dimensiones y cerrar de inmediato
            img_temp = fitz.open(stream=imagen_bytes, filetype=filetype)
            rect = img_temp[0].rect
            img_temp.close()

            # Crear página del tamaño exacto de la imagen para evitar distorsión
            page = doc.new_page(width=rect.width, height=rect.height)
            page.insert_image(rect, stream=imagen_bytes)

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar las imágenes: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/repair', methods=['POST'])
def repair_pdf():
    """
    Intenta recuperar un PDF dañado o corrupto.

    PyMuPDF repara automáticamente errores estructurales al abrir el archivo.
    Se fuerza la lectura completa de cada página para que el motor detecte
    y corrija todos los errores antes de guardar.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF dañado.

    Returns:
        Response: Template con enlace al PDF reparado.

    Parámetros de save():
        garbage=4: limpieza profunda
        deflate=True: compresión
        clean=True: optimización
        linear=True: reorganización para acceso página a página
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_reparado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        # Forzar lectura completa — PyMuPDF repara errores al leer cada página
        # _ indica que el valor retornado no se necesita, solo el efecto de leer
        for page in doc:
            _ = page.get_text()

        doc.save(output_path, garbage=4, deflate=True, clean=True, linear=True)
        doc.close()

    except Exception as e:
        return f'No se pudo reparar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/crop', methods=['POST'])
def crop_pdf():
    """
    Recorta los márgenes de todas las páginas de un PDF.

    Usa CropBox que oculta el contenido fuera del área sin eliminarlo.
    Los márgenes se expresan en porcentaje para funcionar con cualquier
    tamaño de página. Cada margen está limitado a máximo 40%.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a recortar.
        margen_izq, margen_der, margen_sup, margen_inf (int): Porcentaje 0-40.

    Returns:
        Response: Template con enlace al PDF recortado.

    Validaciones:
        - Márgenes opuestos sumando >= 100% -> 400
        - Valores no numéricos -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    try:
        # max/min limita el rango en el servidor aunque el HTML ya lo valide
        margen_izq = max(0, min(40, int(request.form.get('margen_izq', 0))))
        margen_der = max(0, min(40, int(request.form.get('margen_der', 0))))
        margen_sup = max(0, min(40, int(request.form.get('margen_sup', 0))))
        margen_inf = max(0, min(40, int(request.form.get('margen_inf', 0))))
    except ValueError:
        return 'Los márgenes deben ser números enteros.', 400

    # Dos márgenes opuestos pueden sumar más del 100% aunque cada uno sea <= 40%
    if margen_izq + margen_der >= 100 or margen_sup + margen_inf >= 100:
        return 'Los márgenes combinados no pueden superar el 100%.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '_recortado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        for page in doc:
            ancho = page.rect.width
            alto = page.rect.height

            # Calcular el área visible resultante en puntos
            x0 = ancho * (margen_izq / 100)
            y0 = alto * (margen_sup / 100)
            x1 = ancho * (1 - margen_der / 100)
            y1 = alto * (1 - margen_inf / 100)

            page.set_cropbox(fitz.Rect(x0, y0, x1, y1))

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/pdf_to_pdfa', methods=['POST'])
def pdf_to_pdfa():
    """
    Convierte un PDF al formato PDF/A-2B (ISO 19005-2, nivel básico).

    PDF/A es el estándar internacional para archivado a largo plazo.
    Es obligatorio en muchos trámites gubernamentales en Colombia.
    Los metadatos XMP que declaran el cumplimiento del estándar se
    insertan mediante pikepdf que opera a nivel de estructura interna.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a convertir.

    Returns:
        Response: Template con enlace al PDF/A generado.

    Metadatos insertados:
        pdfaid:part = '2' -> versión PDF/A-2
        pdfaid:conformance = 'B' -> nivel básico (reproducción visual)
        dc:format = 'application/pdf' -> tipo de contenido
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_pdfa.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        with pikepdf.open(pdf_path) as pdf:
            # set_pikepdf_as_editor=False evita que pikepdf agregue sus propios metadatos
            with pdf.open_metadata(set_pikepdf_as_editor=False) as meta:
                meta['pdfaid:part'] = '2'
                meta['pdfaid:conformance'] = 'B'
                meta['dc:format'] = 'application/pdf'

            pdf.save(
                output_path,
                compress_streams=True,
                object_stream_mode=pikepdf.ObjectStreamMode.generate
            )

    except Exception as e:
        return f'Error al convertir el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/form_filler', methods=['POST'])
def form_filler():
    """
    Paso 1 del rellenador de formularios: detecta los campos del PDF.

    Lee todos los widgets (campos interactivos) del PDF y los pasa
    al template form_filler.html donde el usuario los puede llenar.
    El nombre del archivo se pasa como campo oculto para el paso 2.

    Parámetros del formulario:
        pdf_file (file): PDF con formulario interactivo.

    Returns:
        Response: Template form_filler.html con los campos detectados.

    Tipos de campo soportados:
        PDF_WIDGET_TYPE_TEXT: campos de texto
        PDF_WIDGET_TYPE_CHECKBOX: casillas de verificación
        PDF_WIDGET_TYPE_LISTBOX: listas desplegables
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    try:
        doc = fitz.open(pdf_path)
        campos = []

        for page_num, page in enumerate(doc):
            for field in page.widgets():
                if field.field_type in [
                    fitz.PDF_WIDGET_TYPE_TEXT,
                    fitz.PDF_WIDGET_TYPE_CHECKBOX,
                    fitz.PDF_WIDGET_TYPE_LISTBOX
                ]:
                    campos.append({
                        'nombre': field.field_name,
                        'tipo': field.field_type,
                        'valor_actual': field.field_value or '',
                        'pagina': page_num
                    })

        doc.close()

        if not campos:
            return 'Este PDF no contiene campos de formulario.', 400

    except Exception as e:
        return f'Error al leer el formulario: {str(e)}', 500

    return render_template('form_filler.html', campos=campos, pdf_nombre=file.filename)


@app.route('/form_filler/guardar', methods=['POST'])
def form_filler_guardar():
    """
    Paso 2 del rellenador de formularios: escribe los valores en el PDF.

    Recibe el nombre del archivo original (pasado como campo oculto desde
    form_filler.html) y los valores de cada campo del formulario.
    Escribe cada valor en su campo correspondiente y guarda el resultado.

    Parámetros del formulario:
        pdf_nombre (str): Nombre del PDF original en /uploads.
        [campo_name] (str/on): Valor de cada campo del formulario.

    Returns:
        Response: Template con enlace al PDF rellenado.

    Validaciones:
        - pdf_nombre vacío -> 400
        - Archivo original no encontrado en /uploads -> 400
    """
    pdf_nombre = request.form.get('pdf_nombre', '')

    if not pdf_nombre:
        return 'Nombre de archivo no encontrado.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, pdf_nombre)

    if not os.path.exists(pdf_path):
        return 'El archivo original ya no está disponible. Sube el PDF nuevamente.', 400

    output_filename = pdf_nombre.rsplit('.', 1)[0] + '_rellenado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        for page in doc:
            for field in page.widgets():
                nombre = field.field_name
                valor = request.form.get(nombre, '')

                # Los checkboxes envían 'on' cuando están marcados, nada cuando no
                if field.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                    field.field_value = valor == 'on'
                else:
                    field.field_value = valor

                field.update()  # aplicar el cambio al PDF en memoria

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al guardar el formulario: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/sign', methods=['POST'])
def sign_pdf():
    """
    Inserta una imagen de firma en una página específica del PDF.

    La firma no es criptográfica — es una imagen visible insertada en
    el área que el usuario defina. La posición y tamaño se expresan
    como porcentaje de las dimensiones de la página.

    Parámetros del formulario:
        pdf_file (file): PDF a firmar.
        firma_file (file): Imagen de la firma (JPG o PNG).
        pagina (int): Número de página (base 1). Por defecto: 1.
        pos_x (int): Posición horizontal en % (0-90). Por defecto: 60.
        pos_y (int): Posición vertical en % (0-90). Por defecto: 80.
        firma_ancho (int): Ancho de la firma en % (5-50). Por defecto: 30.
        firma_alto (int): Alto de la firma en % (5-30). Por defecto: 10.

    Returns:
        Response: Template con enlace al PDF firmado.
    """
    if 'pdf_file' not in request.files or 'firma_file' not in request.files:
        return 'Por favor, sube el PDF y la imagen de tu firma.', 400

    file = request.files['pdf_file']
    firma = request.files['firma_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF válido.', 400

    extensiones_firma = {'.jpg', '.jpeg', '.png'}
    ext_firma = os.path.splitext(firma.filename)[1].lower()
    if ext_firma not in extensiones_firma:
        return 'La firma debe ser una imagen JPG o PNG.', 400

    try:
        pagina_num = max(1, int(request.form.get('pagina', 1)))
        pos_x = max(0, min(90, int(request.form.get('pos_x', 60))))
        pos_y = max(0, min(90, int(request.form.get('pos_y', 80))))
        firma_ancho = max(5, min(50, int(request.form.get('firma_ancho', 30))))
        firma_alto = max(5, min(30, int(request.form.get('firma_alto', 10))))
    except ValueError:
        return 'Los valores de posición deben ser números enteros.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_firmado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        # La validación de página ocurre DESPUÉS de abrir el PDF
        # porque necesitamos page_count para saber si la página existe
        if pagina_num > doc.page_count:
            doc.close()
            return f'El PDF solo tiene {doc.page_count} páginas.', 400

        page = doc[pagina_num - 1]  # convertir de base 1 a base 0
        ancho = page.rect.width
        alto = page.rect.height

        # Calcular el rectángulo donde se insertará la firma
        x0 = ancho * (pos_x / 100)
        y0 = alto * (pos_y / 100)
        x1 = x0 + ancho * (firma_ancho / 100)
        y1 = y0 + alto * (firma_alto / 100)

        firma_bytes = firma.read()
        page.insert_image(fitz.Rect(x0, y0, x1, y1), stream=firma_bytes)

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al firmar el documento: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')

# ─── Nivel Avanzado ───────────────────────────────────────────────────────────

@app.route('/pdf_to_excel', methods=['POST'])
def pdf_to_excel():
    """
    Extrae todas las tablas de un PDF y las exporta a un archivo Excel.

    Cada tabla encontrada se escribe en una hoja separada. El nombre de
    cada hoja indica la página y número de tabla de origen (ej: Pag1_Tabla2).
    Las celdas vacías (None) se convierten a string vacío para evitar
    corrupción del archivo Excel.

    Parámetros del formulario:
        pdf_file (file): PDF con tablas a extraer.

    Returns:
        Response: Template con enlace al archivo Excel generado.

    Validaciones:
        - PDF sin tablas detectables -> 400
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '.xlsx'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Workbook() crea una hoja vacía por defecto — la eliminamos
        tablas_encontradas = 0

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tablas = page.extract_tables()

                for tabla_num, tabla in enumerate(tablas):
                    if not tabla:
                        continue

                    tablas_encontradas += 1
                    nombre_hoja = f"Pag{page_num + 1}_Tabla{tabla_num + 1}"
                    ws = wb.create_sheet(title=nombre_hoja)

                    for fila in tabla:
                        # pdfplumber retorna None en celdas vacías — openpyxl necesita strings
                        fila_limpia = [celda if celda is not None else '' for celda in fila]
                        ws.append(fila_limpia)

        if tablas_encontradas == 0:
            return 'No se encontraron tablas en el PDF.', 400

        wb.save(output_path)

    except Exception as e:
        return f'Error al procesar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/edit', methods=['POST'])
def edit_pdf():
    """
    Agrega texto en una posición específica de una página del PDF.

    La posición se define como porcentaje de las dimensiones de la página
    para funcionar con cualquier tamaño. El color se convierte de hex HTML
    a RGB normalizado con la función auxiliar hex_a_rgb().

    Parámetros del formulario:
        pdf_file (file): PDF a editar.
        texto (str): Texto a insertar.
        pagina (int): Número de página (base 1). Por defecto: 1.
        pos_x (int): Posición horizontal en % (0-95). Por defecto: 10.
        pos_y (int): Posición vertical en % (0-95). Por defecto: 50.
        fontsize (int): Tamaño de fuente en puntos (6-72). Por defecto: 12.
        color (str): Color del texto en hex. Por defecto: '#000000'.

    Returns:
        Response: Template con enlace al PDF editado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    texto = request.form.get('texto', '').strip()
    color_hex = request.form.get('color', '#000000')

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not texto:
        return 'Por favor, escribe el texto a insertar.', 400

    try:
        pagina_num = max(1, int(request.form.get('pagina', 1)))
        pos_x = max(0, min(95, int(request.form.get('pos_x', 10))))
        pos_y = max(0, min(95, int(request.form.get('pos_y', 50))))
        fontsize = max(6, min(72, int(request.form.get('fontsize', 12))))
    except ValueError:
        return 'Los valores de posición deben ser números enteros.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_editado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        if pagina_num > doc.page_count:
            doc.close()
            return f'El PDF solo tiene {doc.page_count} páginas.', 400

        page = doc[pagina_num - 1]
        ancho = page.rect.width
        alto = page.rect.height

        # Convertir posición de porcentaje a puntos
        x = ancho * (pos_x / 100)
        y = alto * (pos_y / 100)

        color = hex_a_rgb(color_hex)  # convertir '#ff0000' a (1.0, 0.0, 0.0)

        page.insert_text(fitz.Point(x, y), texto, fontsize=fontsize, color=color)

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al editar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/censor', methods=['POST'])
def censor_pdf():
    """
    Dibuja rectángulos de color sólido sobre zonas sensibles de una página.

    Las coordenadas se expresan como porcentaje en formato x0,y0,x1,y1
    (una zona por línea). Las líneas malformadas se ignoran silenciosamente
    — solo falla si ninguna zona resulta válida.

    Reutiliza hex_a_rgb() definida en la sección de funciones auxiliares.

    Parámetros del formulario:
        pdf_file (file): PDF a censurar.
        pagina (int): Número de página (base 1). Por defecto: 1.
        zonas (str): Coordenadas en % separadas por líneas. Ej: '10,20,40,30'.
        color (str): Color de censura en hex. Por defecto: '#000000'.

    Returns:
        Response: Template con enlace al PDF censurado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']
    zonas_texto = request.form.get('zonas', '').strip()
    color_hex = request.form.get('color', '#000000')

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    if not zonas_texto:
        return 'Por favor, indica al menos una zona a censurar.', 400

    try:
        pagina_num = max(1, int(request.form.get('pagina', 1)))
    except ValueError:
        return 'El número de página debe ser un entero.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_censurado.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc = fitz.open(pdf_path)

        if pagina_num > doc.page_count:
            doc.close()
            return f'El PDF solo tiene {doc.page_count} páginas.', 400

        page = doc[pagina_num - 1]
        ancho = page.rect.width
        alto = page.rect.height
        color = hex_a_rgb(color_hex)
        zonas_validas = 0

        for linea in zonas_texto.strip().split('\n'):
            partes = linea.strip().split(',')
            if len(partes) != 4:
                continue  # línea malformada — ignorar sin error
            try:
                x0, y0, x1, y1 = [float(p.strip()) for p in partes]
                # Convertir porcentajes a puntos y limitar al rango válido
                x0 = ancho * (max(0, min(100, x0)) / 100)
                y0 = alto  * (max(0, min(100, y0)) / 100)
                x1 = ancho * (max(0, min(100, x1)) / 100)
                y1 = alto  * (max(0, min(100, y1)) / 100)

                # color y fill iguales crean un rectángulo completamente opaco
                page.draw_rect(fitz.Rect(x0, y0, x1, y1), color=color, fill=color)
                zonas_validas += 1
            except ValueError:
                continue  # línea con texto en vez de números — ignorar

        if zonas_validas == 0:
            doc.close()
            return 'No se encontraron zonas válidas. Formato esperado: x0,y0,x1,y1', 400

        doc.save(output_path)
        doc.close()

    except Exception as e:
        return f'Error al censurar el archivo: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/ocr', methods=['POST'])
def ocr_pdf():
    """
    Aplica reconocimiento óptico de caracteres (OCR) a un PDF escaneado.

    Convierte cada página del PDF en imagen, extrae el texto con Tesseract
    y lo incrusta de forma invisible en el PDF resultado. El PDF resultante
    mantiene la apariencia visual original pero el texto es buscable y copiable.

    Requiere en el servidor Ubuntu:
        sudo apt-get install tesseract-ocr tesseract-ocr-spa

    Parámetros del formulario:
        pdf_file (file): PDF escaneado (imágenes sin texto seleccionable).

    Returns:
        Response: Template con enlace al PDF con texto OCR incrustado.

    Truco del texto invisible:
        fontsize=0 + color=(1,1,1) hace el texto blanco sobre blanco —
        invisible visualmente pero indexado por lectores de PDF para búsqueda.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(pdf_path)

    output_filename = file.filename.rsplit('.', 1)[0] + '_ocr.pdf'
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        doc_original = fitz.open(pdf_path)
        doc_nuevo = fitz.open()

        for page in doc_original:
            # Renderizar página a alta resolución para mejor precisión del OCR
            pixmap = page.get_pixmap(dpi=300)
            imagen_bytes = pixmap.tobytes("png")

            # PIL necesita un objeto Image — io.BytesIO convierte bytes a stream
            imagen = Image.open(io.BytesIO(imagen_bytes))
            texto = pytesseract.image_to_string(imagen, lang='spa')

            # Crear página del mismo tamaño en el documento nuevo
            page_nueva = doc_nuevo.new_page(width=page.rect.width, height=page.rect.height)

            # Insertar la imagen original como fondo visual
            page_nueva.insert_image(page_nueva.rect, stream=imagen_bytes)

            # Insertar el texto extraído de forma invisible para que sea buscable
            if texto.strip():
                page_nueva.insert_text(
                    fitz.Point(0, 20),
                    texto,
                    fontsize=0,          # tamaño 0 = invisible
                    color=(1, 1, 1)      # blanco = invisible sobre fondo claro
                )

        doc_original.close()
        doc_nuevo.save(output_path)
        doc_nuevo.close()

    except Exception as e:
        return f'Error al aplicar OCR: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')


@app.route('/compare', methods=['POST'])
def compare_pdf():
    """
    Compara el texto de dos versiones de un PDF y genera un reporte HTML.

    Usa difflib.HtmlDiff para generar una tabla con las diferencias
    resaltadas en color: verde para texto agregado, rojo para eliminado
    y amarillo para cambios. Solo muestra zonas con cambios (context=True)
    más 3 líneas de contexto alrededor de cada diferencia.

    El resultado es un archivo .html porque las diferencias coloreadas
    se visualizan mejor en el navegador que en un PDF estático.

    Parámetros del formulario:
        pdf_file_1 (file): PDF versión original.
        pdf_file_2 (file): PDF versión nueva.

    Returns:
        Response: Template con enlace al reporte HTML de comparación.
    """
    if 'pdf_file_1' not in request.files or 'pdf_file_2' not in request.files:
        return 'Por favor, sube los dos archivos PDF a comparar.', 400

    file1 = request.files['pdf_file_1']
    file2 = request.files['pdf_file_2']

    if file1.filename == '' or not file1.filename.endswith('.pdf'):
        return 'El primer archivo debe ser un PDF válido.', 400

    if file2.filename == '' or not file2.filename.endswith('.pdf'):
        return 'El segundo archivo debe ser un PDF válido.', 400

    # Prefijos para evitar colisión de nombres si los dos archivos se llaman igual
    pdf_path_1 = os.path.join(UPLOAD_FOLDER, 'comparar_1_' + file1.filename)
    pdf_path_2 = os.path.join(UPLOAD_FOLDER, 'comparar_2_' + file2.filename)
    file1.save(pdf_path_1)
    file2.save(pdf_path_2)

    try:
        doc1 = fitz.open(pdf_path_1)
        doc2 = fitz.open(pdf_path_2)

        # Extraer texto línea por línea de todo el documento
        texto1 = []
        texto2 = []
        for page in doc1:
            texto1.extend(page.get_text().splitlines())
        for page in doc2:
            texto2.extend(page.get_text().splitlines())

        doc1.close()
        doc2.close()

        # Generar tabla HTML con diferencias coloreadas
        differ = difflib.HtmlDiff(wrapcolumn=80)
        tabla_html = differ.make_table(
            texto1,
            texto2,
            fromdesc=file1.filename,
            todesc=file2.filename,
            context=True,    # mostrar solo zonas con diferencias
            numlines=3       # más 3 líneas de contexto alrededor de cada cambio
        )

        reporte_html = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Comparación de PDFs</title>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 2rem; background: #f9f9f9; }}
        h1 {{ color: #1a1a2e; font-size: 1.5rem; margin-bottom: 0.5rem; }}
        p {{ color: #555; font-size: 0.9rem; margin-bottom: 1.5rem; }}
        table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
        td {{ padding: 4px 8px; vertical-align: top; white-space: pre-wrap; }}
        .diff_header {{ background: #e8e8e8; font-weight: bold; }}
        .diff_next {{ background: #d0d0d0; }}
        .diff_add {{ background: #aaffaa; }}
        .diff_chg {{ background: #ffff77; }}
        .diff_sub {{ background: #ffaaaa; }}
        th {{ background: #333; color: white; padding: 8px; }}
    </style>
</head>
<body>
    <h1>Reporte de comparación</h1>
    <p>Comparando <strong>{file1.filename}</strong> contra <strong>{file2.filename}</strong></p>
    {tabla_html}
</body>
</html>"""

        output_filename = 'comparacion.html'
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(reporte_html)

    except Exception as e:
        return f'Error al comparar los archivos: {str(e)}', 500

    return render_template('index.html', output_file=f'/download/{output_filename}')

# ─── Ruta heredada ────────────────────────────────────────────────────────────

@app.route('/convert', methods=['POST'])
def convert():
    """
    Convierte un PDF a formato Word (.docx) usando pdf2docx.

    Esta ruta existía antes del desarrollo del proyecto actual y se mantiene
    por compatibilidad. Usa la librería pdf2docx que intenta preservar
    el formato del documento original.

    Parámetros del formulario:
        pdf_file (file): Archivo PDF a convertir.

    Returns:
        Response: Template con enlace al archivo Word generado.
    """
    if 'pdf_file' not in request.files:
        return 'No se ha seleccionado un archivo.', 400

    file = request.files['pdf_file']

    if file.filename == '' or not file.filename.endswith('.pdf'):
        return 'Por favor, suba un archivo PDF.', 400

    filename = secure_filename(file.filename)
    pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(pdf_path)

    output_filename = filename.rsplit('.', 1)[0] + '.docx'
    word_path = os.path.join(OUTPUT_FOLDER, output_filename)

    cv = Converter(pdf_path)
    cv.convert(word_path, start=0, end=None)
    cv.close()

    output_file_url = f'/download/{output_filename}'
    return render_template('index.html', output_file=output_file_url)

# ─── Configuración WSGI ───────────────────────────────────────────────────────

# Variable requerida por mod_wsgi para encontrar la aplicación en el servidor
application = app

# Para pruebas locales descomenta las siguientes líneas:
# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=8080, debug=True)